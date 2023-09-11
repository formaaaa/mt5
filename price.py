from MetaTrader5 import initialize, shutdown, symbol_info_tick
import openpyxl

# Initialize the MetaTrader 5 terminal
if not initialize():
    print("Initialization has failed")
    shutdown()

symbols = ["AAPL.NAS", 	"ABBV.NYSE", 	"ABNB.NAS", 	"ABT.NYSE", 	"ACN.NYSE", 	"ADBE.NAS", 	"ADS.XE",
           "AI.EPA", 	"AIR.EPA", 	"ALV.XE", 	"AMD.NAS", 	"AMZN.NAS", 	"ASML.EAS", 	"AVGO.NAS", 	"AXP.NYSE",
           "AZN.LSE", 	"BA.NYSE", 	"BABA.NYSE", 	"BAC.NYSE", 	"BARC.LSE", 	"BATS.LSE", 	"BAYN.FWB",
           "BBVA.BM", 	"BEI.XE", 	"BHP.ASX", 	"BKNG.NAS", 	"BLK.NYSE", 	"BMW.XE", 	"BMY.NYSE", 	"BNP.EPA",
           "BX.NYSE", 	"CABK.BM", 	"CAT.NYSE", 	"CBA.ASX", 	"CMCSA.NAS", 	"CNHI.MIL", 	"COP.NYSE", 	"COST.NAS",
           "CPTL.SGX", 	"CRM.NYSE", 	"CSCO.NAS", 	"CSL.ASX", 	"CVS.NYSE", 	"CVX.NYSE", 	"DAII.TSE",
           "DHR.NYSE", 	"DIS.NYSE", 	"DKI.TSE", 	"DTE.FWB", 	"DTG.FWB", 	"ENEL.MIL", 	"ENI.MIL", 	"GE.NYSE",
           "GOOGL.NAS", 	"GS.NYSE", 	"HD.NYSE", 	"HEIA.EAS", 	"HIT.TSE", 	"HK.SGX", 	"HMC.NYSE", 	"HON.NAS",
           "HSBA.LSE", 	"IBM.NYSE", 	"INTC.NAS", 	"INTU.NAS", 	"ISP.MIL", 	"ISRG.NAS", 	"ITX.BM",
           "JAR.SGX", 	"JD.NAS", 	"JNJ.NYSE", 	"JPM.NYSE", 	"KEE.TSE", 	"KO.NYSE", 	"LI.NAS", 	"LIN.NYSE",
           "LLY.NYSE", 	"LMT.NYSE", 	"LSEG.LSE", 	"MA.NYSE", 	"MBG.XE", 	"MC.EPA", 	"MCD.NYSE", 	"MDLZ.NAS",
           "META.NAS", 	"MMM.NYSE", 	"MRK.NYSE", 	"MRK.XE", 	"MS.NYSE", 	"MSFT.NAS", 	"MUR.TSE", 	"NAB.ASX",
           "NEE.NYSE", 	"NESTE.OMXH", 	"NFLX.NAS", 	"NID.TSE", 	"NKE.NYSE", 	"NOKIA.OMXH", 	"NOW.NYSE",
           "NVDA.NAS", 	"OL.TSE", 	"OR.EPA", 	"ORCL.NYSE", 	"PEP.NAS", 	"PFE.NYSE", 	"PG.NYSE", 	"PM.NYSE",
           "PYPL.NAS", 	"QCOM.NAS", 	"RACE.MIL", 	"RMS.EPA", 	"RTX.NYSE", 	"SAN.BM", 	"SAP.NYSE",
           "SBUX.NAS", 	"SGX.SGX", 	"SHEL.LSE", 	"SHOP.NYSE", 	"SIAIR.SGX", 	"SIE.XE", 	"SONY.NYSE",
           "STLA.MIL", 	"SVN.TSE", 	"T.NYSE", 	"TCEHY.OTC", 	"TKY.TSE", 	"TM.TSE", 	"TMH.TSE", 	"TMO.NYSE",
           "TMUS.NAS", 	"TSLA.NAS", 	"TTE.EPA", 	"TXN.NAS", 	"UCG.MIL", 	"ULVR.LSE", 	"UNH.NYSE", 	"UNP.NYSE",
           "UPS.NYSE", 	"V.NYSE", 	"VNA.XE", 	"VOD.LSE", 	"VOLVA.SOMX", 	"VOW3.XE", 	"VZ.NYSE", 	"WES.ASX",
           "WFC.NYSE", 	"WLMR.SGX", 	"WMT.NYSE", 	"XOM.NYSE"]

tick_data = {}

for symbol in symbols:
    tick = symbol_info_tick(symbol)
    if tick:
        tick_data[symbol] = tick.ask  # or tick.bid depending on your needs

# Create a new workbook and select the active worksheet
wb = openpyxl.Workbook()
ws = wb.active

# Write data to Excel
row_num = 1
for symbol, price in tick_data.items():
    ws.cell(row=row_num, column=1).value = symbol
    ws.cell(row=row_num, column=2).value = price
    row_num += 1

# Save the workbook
wb.save("real_time_prices.xlsx")

# Shutdown the connection to MetaTrader 5
shutdown()
